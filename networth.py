from sqlalchemy import create_engine, Column, String, Float, Integer, DateTime
from sqlalchemy.orm import declarative_base, sessionmaker
from datetime import datetime, timezone
import pandas as pd

Base = declarative_base()

class AssetValue(Base):
    __tablename__ = 'asset_values'
    
    id = Column(Integer, primary_key=True)
    asset_name = Column(String)
    tag = Column(String)
    value = Column(Float)
    recorded_at = Column(DateTime)

def get_networth_db():
    engine = create_engine('sqlite:///budget_bot.db')
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine)
    return Session()

def import_from_worthit(filepath):
    import openpyxl
    import ast

    wb = openpyxl.load_workbook(filepath, read_only=True)
    db = get_networth_db()
    imported = 0
    skipped_sheets = []

    # Skip the overview sheet
    asset_sheets = [s for s in wb.sheetnames if s != 'Overview']

    for sheet_name in asset_sheets:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        # Parse key fields from sheet
        data = {}
        for row in rows:
            if row[0] in ('name', 'tag', 'is_archived', 'values', 'values_market'):
                data[row[0]] = row[1]

        # Skip archived assets
        if data.get('is_archived') == 'true':
            skipped_sheets.append(data.get('name', sheet_name))
            continue

        asset_name = data.get('name', sheet_name)
        tag = data.get('tag', 'Other')
        # Clean tag of emojis for cleaner display
        tag = tag.encode('ascii', 'ignore').decode('ascii').strip()
        if not tag:
            tag = 'Other'

        # Parse values - combine manual and market values
        all_values = {}
        for key in ('values', 'values_market'):
            raw = data.get(key, '{}')
            if raw and raw != '{}':
                try:
                    parsed = ast.literal_eval(str(raw))
                    all_values.update(parsed)
                except:
                    pass

        # Convert timestamps and save
        for ts_ms, value in all_values.items():
            if not value:
                continue
            recorded_at = datetime.fromtimestamp(int(ts_ms) / 1000, tz=timezone.utc).replace(tzinfo=None)

            # Check for duplicate
            exists = db.query(AssetValue).filter_by(
                asset_name=asset_name,
                recorded_at=recorded_at
            ).first()

            if not exists:
                db.add(AssetValue(
                    asset_name=asset_name,
                    tag=tag,
                    value=float(value),
                    recorded_at=recorded_at
                ))
                imported += 1

    db.commit()
    db.close()
    return imported, skipped_sheets

def get_all_asset_history():
    db = get_networth_db()
    records = db.query(AssetValue).order_by(AssetValue.recorded_at).all()
    db.close()
    return records

def get_asset_names():
    db = get_networth_db()
    names = db.query(AssetValue.asset_name).distinct().all()
    db.close()
    return [n[0] for n in names]

def get_asset_name_tags():
    """Returns {asset_name: tag} - lets callers group assets by category
    (e.g. all 'Pension'-tagged assets) without the caller needing to know
    the exact set of asset names in advance."""
    db = get_networth_db()
    rows = db.query(AssetValue.asset_name, AssetValue.tag).distinct().all()
    db.close()
    return dict(rows)

def get_total_net_worth_series(asset_names=None):
    """Returns a DataFrame ['Date', 'Total']: one row per calendar day any
    asset had a recorded value, across the FULL history (never filtered),
    with each asset forward-filled to its last known value so a day only one
    asset was updated still counts every other asset's last value.

    If asset_names is given, only those assets are summed - e.g. to project
    "just my pensions and ISA" rather than everything tracked."""
    records = get_all_asset_history()
    if asset_names is not None:
        records = [r for r in records if r.asset_name in asset_names]
    if not records:
        return pd.DataFrame(columns=['Date', 'Total'])

    df = pd.DataFrame([{
        'Date': r.recorded_at, 'Asset': r.asset_name, 'Value': r.value
    } for r in records])
    df['DateOnly'] = df['Date'].dt.date
    asset_pivot = df.sort_values('Date').pivot_table(
        index='DateOnly', columns='Asset', values='Value', aggfunc='last'
    ).ffill()
    daily_total = asset_pivot.sum(axis=1).reset_index()
    daily_total.columns = ['Date', 'Total']
    return daily_total

# Minimum span of history before an annualised growth rate is meaningful -
# a few days of data would produce a wildly extrapolated rate
MIN_GROWTH_RATE_DAYS = 90

def calculate_growth_rate(series=None):
    """Computes a simple annualised growth rate from the first to last point
    of a total net worth series (as from get_total_net_worth_series()). This
    is a BLENDED rate - it reflects market movement AND the user's own
    contributions/withdrawals, not a pure investment return.

    Returns (rate, span_days). rate is None when there are fewer than 2 data
    points, the starting value isn't positive, or the span is too short to
    annualise meaningfully."""
    if series is None:
        series = get_total_net_worth_series()

    if len(series) < 2:
        return None, 0

    first_row, last_row = series.iloc[0], series.iloc[-1]
    first_value, last_value = first_row['Total'], last_row['Total']
    span_days = (pd.Timestamp(last_row['Date']) - pd.Timestamp(first_row['Date'])).days

    if first_value <= 0 or span_days < MIN_GROWTH_RATE_DAYS:
        return None, span_days

    years = span_days / 365.25
    rate = (last_value / first_value) ** (1 / years) - 1
    return rate, span_days

def project_net_worth(years=5, asset_names=None, series=None):
    """Projects net worth forward `years` years from the latest known data
    point, using calculate_growth_rate(). If asset_names is given, only
    projects that subset (e.g. just pensions + ISA) rather than everything.
    Returns a dict:
      ok, reason ('no_data' | 'insufficient_history' | None),
      anchor_date, anchor_value, projected_value, rate, span_days, years,
      start_date (the actual first data point the growth rate was computed
      from - NOT the same as anchor_date, which is the most recent point the
      projection starts FROM)"""
    if series is None:
        series = get_total_net_worth_series(asset_names=asset_names)

    if series.empty:
        return {'ok': False, 'reason': 'no_data', 'anchor_date': None,
                'anchor_value': None, 'projected_value': None,
                'rate': None, 'span_days': 0, 'years': years, 'start_date': None}

    rate, span_days = calculate_growth_rate(series)
    anchor_row = series.iloc[-1]
    anchor_date, anchor_value = anchor_row['Date'], anchor_row['Total']
    start_date = series.iloc[0]['Date']

    if rate is None:
        return {'ok': False, 'reason': 'insufficient_history', 'anchor_date': anchor_date,
                'anchor_value': anchor_value, 'projected_value': None,
                'rate': None, 'span_days': span_days, 'years': years, 'start_date': start_date}

    projected_value = anchor_value * (1 + rate) ** years
    return {'ok': True, 'reason': None, 'anchor_date': anchor_date,
            'anchor_value': anchor_value, 'projected_value': projected_value,
            'rate': rate, 'span_days': span_days, 'years': years, 'start_date': start_date}

def update_asset_value(asset_name, value, tag=None):
    db = get_networth_db()

    # Get tag from existing records if not provided
    if not tag:
        existing = db.query(AssetValue).filter_by(
            asset_name=asset_name
        ).first()
        tag = existing.tag if existing else 'Other'

    db.add(AssetValue(
        asset_name=asset_name,
        tag=tag,
        value=float(value),
        recorded_at=datetime.utcnow()
    ))
    db.commit()
    db.close()

def delete_asset(asset_name):
    db = get_networth_db()
    db.query(AssetValue).filter_by(asset_name=asset_name).delete()
    db.commit()
    db.close()

def delete_asset_value(record_id):
    db = get_networth_db()
    db.query(AssetValue).filter_by(id=record_id).delete()
    db.commit()
    db.close()

def get_asset_tags():
    db = get_networth_db()
    tags = db.query(AssetValue.tag).distinct().all()
    db.close()
    return sorted(t[0] for t in tags if t[0])